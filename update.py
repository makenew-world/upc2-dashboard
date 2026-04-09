#!/usr/bin/env python3
"""
UPC2 Dashboard — Daily Data Updater
=====================================
วิธีใช้:
  python3 update.py                          # หา SD0002*.xlsx ล่าสุดใน ~/Downloads อัตโนมัติ
  python3 update.py /path/to/SD0002_file.xlsx  # ระบุไฟล์เอง

ผลลัพธ์: เขียน data.json ในโฟลเดอร์นี้
"""

import sys, os, json, glob, re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌  กรุณาติดตั้ง openpyxl ก่อน:  pip3 install openpyxl")
    sys.exit(1)

THAI_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
               'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

# ── Targets per month (key = "MM") ─────────────────────────────────────────
MONTHLY_TARGETS = {
    "04": {
        "PU4": {"ESPOGEN":1351000,"EPOTIV":374500,"EUVAX":39165,"ZEMIGLO":461250,"ZEMIMET":4875,"ZEMIDAPA":48000},
        "PU5": {"ESPOGEN":2394000,"EPOTIV":1523900,"EUVAX":38024,"ZEMIGLO":354000,"ZEMIMET":5250,"ZEMIDAPA":32000},
        "PU6": {"ESPOGEN":1183000,"EPOTIV":339500,"EUVAX":16956.38,"ZEMIGLO":253500,"ZEMIMET":4875,"ZEMIDAPA":20000},
        "DU3": {"ZEMIGLO":2096250,"ZEMIMET":71250,"ZEMIDAPA":60000},
        "DU4": {"ZEMIGLO":1845000,"ZEMIMET":26250,"ZEMIDAPA":60000},
    },
    "05": {
        "PU4": {"ESPOGEN":1737000,"EPOTIV":481500,"EUVAX":50355,"ZEMIGLO":522750,"ZEMIMET":5525,"ZEMIDAPA":60000},
        "PU5": {"ESPOGEN":3078000,"EPOTIV":1959300,"EUVAX":48888,"ZEMIGLO":401200,"ZEMIMET":5950,"ZEMIDAPA":40000},
        "PU6": {"ESPOGEN":1521000,"EPOTIV":436500,"EUVAX":21801.06,"ZEMIGLO":287300,"ZEMIMET":5525,"ZEMIDAPA":25000},
        "DU3": {"ZEMIGLO":2375750,"ZEMIMET":80750,"ZEMIDAPA":75000},
        "DU4": {"ZEMIGLO":2091000,"ZEMIMET":29750,"ZEMIDAPA":75000},
    },
    "06": {
        "PU4": {"ESPOGEN":1737000,"EPOTIV":481500,"EUVAX":50355,"ZEMIGLO":553500,"ZEMIMET":5850,"ZEMIDAPA":72000},
        "PU5": {"ESPOGEN":3078000,"EPOTIV":1959300,"EUVAX":48888,"ZEMIGLO":424800,"ZEMIMET":6300,"ZEMIDAPA":48000},
        "PU6": {"ESPOGEN":1521000,"EPOTIV":436500,"EUVAX":21801.06,"ZEMIGLO":304200,"ZEMIMET":5850,"ZEMIDAPA":30000},
        "DU3": {"ZEMIGLO":2515500,"ZEMIMET":85500,"ZEMIDAPA":90000},
        "DU4": {"ZEMIGLO":2214000,"ZEMIMET":31500,"ZEMIDAPA":90000},
    },
}

# ── Scheme definitions per month ───────────────────────────────────────────
def build_scheme_def(tgt):
    """Build SCHEME_DEF from a monthly target dict."""
    def epo(a):  return (tgt[a].get("ESPOGEN",0) + tgt[a].get("EPOTIV",0))
    def zemi(a): return (tgt[a].get("ZEMIGLO",0) + tgt[a].get("ZEMIMET",0) + tgt[a].get("ZEMIDAPA",0))
    def total(a):return sum(tgt[a].values())
    defs = {}
    for a in ["PU4","PU5","PU6"]:
        defs[a] = [
            {"name":"EPO Family","brands":["ESPOGEN","EPOTIV"],"tgt":epo(a)},
            {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":zemi(a)},
            {"name":"TOTAL","brands":None,"tgt":total(a)},
        ]
    for a in ["DU3","DU4"]:
        defs[a] = [
            {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":zemi(a)},
            {"name":"Zemidapa","brands":["ZEMIDAPA"],"tgt":tgt[a].get("ZEMIDAPA",0)},
        ]
    defs["MGR"] = [
        {"name":"EPO Family","brands":["ESPOGEN","EPOTIV"],"tgt":sum(epo(a) for a in ["PU4","PU5","PU6"])},
        {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":sum(zemi(a) for a in ["PU4","PU5","PU6","DU3","DU4"])},
        {"name":"TOTAL","brands":None,"tgt":sum(total(a) for a in ["PU4","PU5","PU6"]) + sum(zemi(a) for a in ["DU3","DU4"])},
    ]
    return defs

# ── Q2 scheme definition (Apr–Jun) ─────────────────────────────────────────
def build_q2_scheme(prev_act_by_area=None):
    """Build Q2_SCHEME. prev_act_by_area = dict of {area: {scheme_name: actual}} for months already passed."""
    months = ["04","05","06"]
    prev_act_by_area = prev_act_by_area or {}

    def q2_total(area, brands_or_none):
        total = 0
        for mm in months:
            t = MONTHLY_TARGETS.get(mm, {}).get(area, {})
            if brands_or_none is None:
                total += sum(t.values())
            else:
                total += sum(t.get(b,0) for b in brands_or_none)
        return total

    defs = {}
    for a in ["PU4","PU5","PU6"]:
        prev = prev_act_by_area.get(a, {})
        defs[a] = [
            {"name":"EPO Family","brands":["ESPOGEN","EPOTIV"],"tgt":q2_total(a,["ESPOGEN","EPOTIV"]),"janFebAct":prev.get("EPO Family",0)},
            {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":q2_total(a,["ZEMIGLO","ZEMIMET","ZEMIDAPA"]),"janFebAct":prev.get("ZEMI Family",0)},
            {"name":"TOTAL","brands":None,"tgt":q2_total(a,None),"janFebAct":prev.get("TOTAL",0)},
        ]
    for a in ["DU3","DU4"]:
        prev = prev_act_by_area.get(a, {})
        defs[a] = [
            {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":q2_total(a,["ZEMIGLO","ZEMIMET","ZEMIDAPA"]),"janFebAct":prev.get("ZEMI Family",0)},
            {"name":"Zemidapa","brands":["ZEMIDAPA"],"tgt":q2_total(a,["ZEMIDAPA"]),"janFebAct":prev.get("Zemidapa",0)},
        ]
    # MGR
    prev = prev_act_by_area.get("MGR", {})
    pu_areas = ["PU4","PU5","PU6"]
    all_areas = ["PU4","PU5","PU6","DU3","DU4"]
    defs["MGR"] = [
        {"name":"EPO Family","brands":["ESPOGEN","EPOTIV"],"tgt":sum(q2_total(a,["ESPOGEN","EPOTIV"]) for a in pu_areas),"janFebAct":prev.get("EPO Family",0)},
        {"name":"ZEMI Family","brands":["ZEMIGLO","ZEMIMET","ZEMIDAPA"],"tgt":sum(q2_total(a,["ZEMIGLO","ZEMIMET","ZEMIDAPA"]) for a in all_areas),"janFebAct":prev.get("ZEMI Family",0)},
        {"name":"TOTAL","brands":None,"tgt":sum(q2_total(a,None) for a in pu_areas)+sum(q2_total(a,["ZEMIGLO","ZEMIMET","ZEMIDAPA"]) for a in ["DU3","DU4"]),"janFebAct":prev.get("TOTAL",0)},
    ]
    return defs

# ── Find Excel file ─────────────────────────────────────────────────────────
def find_excel():
    if len(sys.argv) > 1:
        path = sys.argv[1]
        if not os.path.exists(path):
            print(f"❌  ไม่พบไฟล์: {path}")
            sys.exit(1)
        return path
    pattern = os.path.expanduser("~/Downloads/SD0002*.xlsx")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        print("❌  ไม่พบ SD0002*.xlsx ใน ~/Downloads")
        print("    กรุณาระบุ path:  python3 update.py /path/to/file.xlsx")
        sys.exit(1)
    print(f"📂  ใช้ไฟล์: {os.path.basename(files[0])}")
    return files[0]

# ── Parse Excel ─────────────────────────────────────────────────────────────
def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Raw Data" not in wb.sheetnames:
        print("❌  ไม่พบ sheet 'Raw Data' ในไฟล์")
        sys.exit(1)
    ws = wb["Raw Data"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌  Raw Data ว่างเปล่า")
        sys.exit(1)

    entries = []
    latest_date = None
    detected_month = None

    for row in rows[1:]:
        billing_date = row[1]
        if not isinstance(billing_date, datetime):
            continue
        area     = row[2]
        customer = (row[4] or "").replace('"', '\\"')
        brand    = row[5]
        qty      = int(row[8] or 0)
        amount   = float(row[10] or 0)
        borrow   = float(row[40] or 0)
        value    = amount - borrow
        if value == 0:
            continue
        if not area or not brand:
            continue

        date_str = billing_date.strftime("%m-%d")
        v_fmt = int(value) if value == int(value) else round(value, 2)
        entries.append({"d": date_str, "a": area, "c": customer, "b": brand, "q": qty, "v": v_fmt})

        if latest_date is None or billing_date > latest_date:
            latest_date = billing_date
        if detected_month is None:
            detected_month = billing_date.strftime("%m")

    entries.sort(key=lambda x: x["d"])
    return entries, latest_date, detected_month

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("🚀  UPC2 Dashboard — Daily Updater")
    print("─" * 40)

    excel_path = find_excel()
    entries, latest_date, month_str = parse_excel(excel_path)

    if not entries:
        print("⚠️   ไม่มีข้อมูลใน Raw Data (ทุกแถวมียอด 0)")
        sys.exit(1)

    # Labels
    thai_year  = latest_date.year + 543
    mtd_label  = f"{THAI_MONTHS[latest_date.month]} {thai_year}"
    data_date  = f"{latest_date.day} {THAI_MONTHS[latest_date.month]} {thai_year}"

    # Targets
    tgt = MONTHLY_TARGETS.get(month_str)
    if not tgt:
        print(f"⚠️   ไม่มี target สำหรับเดือน {month_str} — กรุณาเพิ่มใน MONTHLY_TARGETS ใน update.py")
        tgt = {}

    scheme_def = build_scheme_def(tgt) if tgt else {}
    q2_scheme  = build_q2_scheme()

    # Build output
    data = {
        "dataDate":  data_date,
        "mtdLabel":  mtd_label,
        "raw":       entries,
        "aprTgt":    tgt,
        "schemeDef": scheme_def,
        "q2Scheme":  q2_scheme,
    }

    # Write data.json next to this script
    out_dir  = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    # Also copy to /tmp/upc2_dashboard/ for local preview
    tmp_dir = "/tmp/upc2_dashboard"
    if os.path.isdir(tmp_dir):
        import shutil
        shutil.copy(out_path, os.path.join(tmp_dir, "data.json"))
        shutil.copy(os.path.join(out_dir, "UPC2_Dashboard_v8_publish.jsx"),
                    os.path.join(tmp_dir, "UPC2_Dashboard_v8_publish.jsx"))
        print(f"🔄  ซิงค์ไปยัง /tmp/upc2_dashboard/ แล้ว")

    print(f"\n✅  สำเร็จ!")
    print(f"   📅  วันที่ล่าสุด : {data_date}")
    print(f"   📊  เดือน        : {mtd_label}")
    print(f"   📋  จำนวนรายการ  : {len(entries)} รายการ")
    print(f"   💾  บันทึกไปที่  : {out_path}")
    print(f"\n   ถ้าใช้ GitHub Pages: git add data.json && git commit -m 'data: {data_date}' && git push")

if __name__ == "__main__":
    main()
